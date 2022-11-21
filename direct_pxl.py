import re
from openpyxl.styles import Font, Color, PatternFill
from pxl import Exel_work
import os
import openpyxl

class Operation:
    def __init__(self, wb_path: str,):
        self.wb_path = wb_path
        self.wb = openpyxl.load_workbook(wb_path)
        self.sh = self.wb.worksheets[0]


    def get_list_from_sh_column(self, *columns: str, start_from_row: int = 1, reference_column: str = 'A',
                                del_last_empty_rows=False):
        """
        Если нужно получить значения определенного столбца в файле exel. Предварительно отрезуются пустые строки снизу.


        :param wb_path: путь до файла, в котором итерируются ячейки
        :param column: столбец, в котором итерируются ячейки
        :param start_from_row: начало итерации для удаления пустых строк
        :param reference_column: показательный столбец, по которому отсчитывается количество строк, то есть,
            не пустое значение ячейки этого столбца гарантирует, что вся строка подлежит отработке. По умолчанию "А"
        :return: Возвращает список значений столбца в ячейке exel
        """
        list = []
        if del_last_empty_rows is True:
            Exel_work().delete_last_empty_rows(self.wb_path, start_from_row)

        for n, row in enumerate(self.sh[reference_column]):
            if n + 1 < start_from_row:
                continue
            corteg = []
            for column in columns:
                corteg.append(self.sh[f'{column}{n + 1}'].value)
            list.append(tuple(corteg))
        return list

    def get_phrase_and_question(self) -> list:
        phrase = self.sh['A1'].value
        question = self.sh['A2'].value
        return [phrase, question]

    def change_value_in_cell(self, row, column, value, saving: bool = True):
        self.sh.cell(row=row, column=column, value=value)
        if saving is True:
            self.save_document()

    def mark_cell(self, row: int, column: int, color: str = "ffff00", saving: bool = True):
        self.sh.cell(row=row, column=column).fill = PatternFill('solid', fgColor=color)
        if saving is True:
            self.save_document()

    def save_document(self):
        self.wb.save(self.wb_path)

    def sort_plan_by_dirs(self, path_files, path_file_lib):

        """
        ТРЕБУЕТСЯ ДОРАБОТКА!!!
        Сортирует файлы планов проверок по папкам. Необходима папка на работем столе "выгрузка файлов по областям", а так же не забудь в объявленных переменных изменить полный путь
        path_files и  path_file_lib соответствено
        :return:
        """
        Exel_work().sort_from_lib(path_files, path_file_lib)

    def merge_cells(self, start_with_row: int = 2):
        column_B = self.sh['B']
        max_row = self.sh.max_row

        for number_row, cell in enumerate(column_B):
            print(number_row)
            if number_row < start_with_row-1:
                continue
            merged_cell_current = self.sh.cell(column=3, row=number_row+1).value


            if cell.value is None and merged_cell_current is not None:

                merged_cell_past = self.sh.cell(column=3, row=number_row).value
                # print(f'строка {number_row+1} будем объединять \n{merged_cell_past} \nс \n{merged_cell_current}')
                self.sh.cell(column=3, row=number_row, value=f'{merged_cell_past} ({merged_cell_current})')
                self.sh.cell(column=3, row=number_row+1, value='')
                # self.sh.delete_rows(number_row+1, amount=1)




def main():
    operation = Operation('/Volumes/KINGSTON/построчные таблицы/табл 9.1.xlsx')
    operation.merge_cells(
        start_with_row=4
    )
    operation.save_document()



if __name__ == '__main__':
    main()


# dire = '/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/Межрегиональное РПН Республике Крым и городу федерального значения Севастополю'
# Exel_work().coint_knm(dire, 18)
# Exel_work().coint_deyatel(dire, 18)

# dir_TU = '/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/Межрегиональное РПН Республике Крым и городу федерального значения Севастополю'
#
# Exel_work().count_risk(dir_TU, 18)


# dir = '/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/РПН железнодорожному транспорту'
# Exel_work().sort_from_lib()